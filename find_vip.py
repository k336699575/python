import openpyxl

  
wb = openpyxl.load_workbook('names_and_savings.xlsx')

ws = wb['工作表1']

num = 1

for row in ws.iter_rows():
    money: int =row[2].value
    if money >= 20000:
        ws.cell(row=num, column=4).value ="VIP"
    num += 1    
# print(ws.cell(row=1, column=1).value)
# print(ws.cell(row=1, column=2).value)
# print(ws.cell(row=1, column=3).value)

      
    wb.save('names_and_savings_new.xlsx')
